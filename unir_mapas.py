import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import requests
import re
import os
from xml.etree import ElementTree as ET
from openpyxl import load_workbook

os.environ.pop("REQUESTS_CA_BUNDLE", None)
os.environ.pop("SSL_CERT_FILE", None)

NS = "http://www.opengis.net/kml/2.2"
ET.register_namespace("", NS)

# Colores KML en formato AABBGGRR (formato KML)
COLORES = [
    ("Rojo",     "ff0000ff"),
    ("Azul",     "ffff0000"),
    ("Verde",    "ff00ff00"),
    ("Amarillo", "ff00ffff"),
    ("Naranja",  "ff0080ff"),
    ("Morado",   "ffff00ff"),
    ("Celeste",  "ffffff00"),
    ("Rosa",     "ff8080ff"),
]

mapas = []  # lista de (nombre, url, color_kml)

# ── GUI ──────────────────────────────────────────────────────────────────────

root = tk.Tk()
root.title("Unir Mapas KML")
root.resizable(False, False)

frame_top = tk.Frame(root, padx=10, pady=10)
frame_top.pack(fill="x")

tk.Label(frame_top, text="Nombre del mapa:").grid(row=0, column=0, sticky="w")
entry_nombre = tk.Entry(frame_top, width=35)
entry_nombre.grid(row=0, column=1, padx=5, pady=3)

tk.Label(frame_top, text="Enlace Google Maps:").grid(row=1, column=0, sticky="w")
entry_url = tk.Entry(frame_top, width=35)
entry_url.grid(row=1, column=1, padx=5, pady=3)

tk.Label(frame_top, text="Color:").grid(row=2, column=0, sticky="w")
color_var = tk.StringVar(value=COLORES[0][0])
combo_color = ttk.Combobox(frame_top, textvariable=color_var,
                           values=[c[0] for c in COLORES], state="readonly", width=15)
combo_color.grid(row=2, column=1, sticky="w", padx=5, pady=3)

# ── Lista de mapas agregados ──────────────────────────────────────────────────

frame_list = tk.Frame(root, padx=10)
frame_list.pack(fill="both", expand=True)

cols = ("Nombre", "Color", "Enlace")
tabla = ttk.Treeview(frame_list, columns=cols, show="headings", height=8)
for c in cols:
    tabla.heading(c, text=c)
tabla.column("Nombre", width=120)
tabla.column("Color",  width=70)
tabla.column("Enlace", width=300)
tabla.pack(side="left", fill="both", expand=True)

scroll = ttk.Scrollbar(frame_list, orient="vertical", command=tabla.yview)
tabla.configure(yscrollcommand=scroll.set)
scroll.pack(side="right", fill="y")

# ── Botones ───────────────────────────────────────────────────────────────────

def color_kml(nombre_color):
    for n, k in COLORES:
        if n == nombre_color:
            return k
    return COLORES[0][1]

def agregar():
    nombre = entry_nombre.get().strip()
    url    = entry_url.get().strip()
    color  = color_var.get()
    if not nombre or not url:
        messagebox.showwarning("Faltan datos", "Ingresa nombre y enlace.")
        return
    if not re.search(r"mid=([^&]+)", url):
        messagebox.showerror("Enlace inválido", "No se encontró el ID del mapa en el enlace.")
        return
    mapas.append((nombre, url, color_kml(color)))
    tabla.insert("", "end", values=(nombre, color, url))
    entry_nombre.delete(0, "end")
    entry_url.delete(0, "end")

def eliminar():
    sel = tabla.selection()
    if not sel:
        return
    idx = tabla.index(sel[0])
    tabla.delete(sel[0])
    mapas.pop(idx)

def hacer_estilo_inline(ckml):
    """Crea un elemento <Style> con estilos inline para My Maps."""
    style = ET.Element(f"{{{NS}}}Style")

    # Icono con color
    icon_style = ET.SubElement(style, f"{{{NS}}}IconStyle")
    ET.SubElement(icon_style, f"{{{NS}}}color").text = ckml
    ET.SubElement(icon_style, f"{{{NS}}}scale").text = "1.0"
    icon = ET.SubElement(icon_style, f"{{{NS}}}Icon")
    ET.SubElement(icon, f"{{{NS}}}href").text = "https://maps.google.com/mapfiles/kml/paddle/wht-blank.png"

    # Etiqueta con color
    label_style = ET.SubElement(style, f"{{{NS}}}LabelStyle")
    ET.SubElement(label_style, f"{{{NS}}}color").text = ckml
    ET.SubElement(label_style, f"{{{NS}}}scale").text = "0.8"

    # Línea con color
    line_style = ET.SubElement(style, f"{{{NS}}}LineStyle")
    ET.SubElement(line_style, f"{{{NS}}}color").text = ckml
    ET.SubElement(line_style, f"{{{NS}}}width").text = "3"

    # Polígono semitransparente
    poly_style = ET.SubElement(style, f"{{{NS}}}PolyStyle")
    ET.SubElement(poly_style, f"{{{NS}}}color").text = "88" + ckml[2:]
    ET.SubElement(poly_style, f"{{{NS}}}fill").text = "1"
    ET.SubElement(poly_style, f"{{{NS}}}outline").text = "1"

    return style


def importar_excel():
    archivo = filedialog.askopenfilename(
        title="Selecciona el Excel de vendedores",
        filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
    )
    if not archivo:
        return

    try:
        wb = load_workbook(archivo)
        ws = wb.active
        
        # Limpiar tabla y lista
        for item in tabla.get_children():
            tabla.delete(item)
        mapas.clear()

        # Leer desde fila 2 (saltar encabezado)
        contador = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:  # nombre y enlace
                continue
            
            nombre = str(row[0]).strip()
            enlace = str(row[1]).strip()
            
            # Detectar separador
            if "SEPARADOR" in nombre.upper():
                continue
            
            if not re.search(r"mid=([^&]+)", enlace):
                continue
            
            # Asignar color rotativo
            color_kml = COLORES[contador % len(COLORES)][1]
            color_nombre = COLORES[contador % len(COLORES)][0]
            
            mapas.append((nombre, enlace, color_kml))
            tabla.insert("", "end", values=(nombre, color_nombre, enlace))
            contador += 1

        messagebox.showinfo("Importado", f"Se importaron {contador} mapas del Excel.")
    
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo leer el Excel:\n{e}")


def importar_por_sesiones():
    archivo = filedialog.askopenfilename(
        title="Selecciona el Excel de vendedores",
        filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")]
    )
    if not archivo:
        return

    try:
        wb = load_workbook(archivo)
        ws = wb.active
        
        sesiones = []
        sesion_actual = []
        
        # Leer y separar por sesiones
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]:
                continue
            
            nombre = str(row[0]).strip()
            enlace = str(row[1]).strip()
            
            # Detectar separador
            if "SEPARADOR" in nombre.upper():
                if sesion_actual:
                    sesiones.append(sesion_actual)
                    sesion_actual = []
                continue
            
            if not re.search(r"mid=([^&]+)", enlace):
                continue
            
            sesion_actual.append((nombre, enlace))
        
        # Agregar última sesión
        if sesion_actual:
            sesiones.append(sesion_actual)
        
        if not sesiones:
            messagebox.showwarning("Sin datos", "No se encontraron mapas en el Excel.")
            return
        
        # Generar un KML por sesión
        for idx, sesion in enumerate(sesiones, 1):
            generar_kml_sesion(sesion, idx)
        
        messagebox.showinfo("Completado", f"✔ Se generaron {len(sesiones)} archivos KML:\n" + 
                           "\n".join([f"mapa_sesion_{i}.kml" for i in range(1, len(sesiones)+1)]))
    
    except Exception as e:
        messagebox.showerror("Error", f"No se pudo procesar:\n{e}")


def generar_kml_sesion(sesion, numero):
    root_kml = ET.Element(f"{{{NS}}}kml")
    doc = ET.SubElement(root_kml, f"{{{NS}}}Document")
    ET.SubElement(doc, f"{{{NS}}}name").text = f"Mapa Sesión {numero}"

    for idx, (nombre, url) in enumerate(sesion):
        match = re.search(r"mid=([^&]+)", url)
        if not match:
            continue
        
        mid = match.group(1)
        kml_url = f"https://www.google.com/maps/d/kml?mid={mid}&forcekml=1"
        color_kml = COLORES[idx % len(COLORES)][1]

        try:
            headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)'}
            r = requests.get(kml_url, verify=False, timeout=60, headers=headers)
            
            if r.status_code != 200:
                continue

            tree = ET.fromstring(r.content)
            src_doc = tree.find(f".//{{{NS}}}Document") or tree

            folder = ET.SubElement(doc, f"{{{NS}}}Folder")
            ET.SubElement(folder, f"{{{NS}}}name").text = nombre

            for elem in src_doc.iter(f"{{{NS}}}Placemark"):
                name_el = elem.find(f"{{{NS}}}name")
                if name_el is None:
                    name_el = ET.SubElement(elem, f"{{{NS}}}name")
                    name_el.text = nombre

                for tag in (f"{{{NS}}}styleUrl", f"{{{NS}}}Style", f"{{{NS}}}StyleMap"):
                    for old in elem.findall(tag):
                        elem.remove(old)

                elem.insert(0, hacer_estilo_inline(color_kml))
                folder.append(elem)

        except:
            pass

    tree_out = ET.ElementTree(root_kml)
    ET.indent(tree_out, space="  ")
    with open(f"mapa_sesion_{numero}.kml", "wb") as f:
        tree_out.write(f, xml_declaration=True, encoding="utf-8")


def generar():
    if not mapas:
        messagebox.showwarning("Sin mapas", "Agrega al menos un mapa.")
        return

    root_kml = ET.Element(f"{{{NS}}}kml")
    doc = ET.SubElement(root_kml, f"{{{NS}}}Document")
    ET.SubElement(doc, f"{{{NS}}}name").text = "Mapa Unido"

    errores = []
    exitosos = 0
    
    for idx, (nombre, url, ckml) in enumerate(mapas, 1):
        match = re.search(r"mid=([^&]+)", url)
        if not match:
            continue
        mid = match.group(1)
        kml_url = f"https://www.google.com/maps/d/kml?mid={mid}&forcekml=1"

        try:
            # Agregar headers y aumentar timeout
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
            }
            r = requests.get(kml_url, verify=False, timeout=60, headers=headers)
            
            if r.status_code != 200:
                errores.append(f"{nombre}: HTTP {r.status_code}")
                continue

            tree = ET.fromstring(r.content)
            src_doc = tree.find(f".//{{{NS}}}Document") or tree

            folder = ET.SubElement(doc, f"{{{NS}}}Folder")
            ET.SubElement(folder, f"{{{NS}}}name").text = nombre

            for elem in src_doc.iter(f"{{{NS}}}Placemark"):
                # Conservar nombre original del placemark
                name_el = elem.find(f"{{{NS}}}name")
                if name_el is None:
                    name_el = ET.SubElement(elem, f"{{{NS}}}name")
                    name_el.text = nombre

                # Eliminar estilos anteriores (styleUrl y Style)
                for tag in (f"{{{NS}}}styleUrl", f"{{{NS}}}Style", f"{{{NS}}}StyleMap"):
                    for old in elem.findall(tag):
                        elem.remove(old)

                # Insertar estilo inline al inicio del Placemark
                elem.insert(0, hacer_estilo_inline(ckml))

                folder.append(elem)
            
            exitosos += 1
            print(f"[{idx}/{len(mapas)}] ✓ {nombre}")

        except requests.exceptions.Timeout:
            errores.append(f"{nombre}: Timeout (tardó mucho)")
        except requests.exceptions.ConnectionError:
            errores.append(f"{nombre}: Error de conexión")
        except Exception as e:
            errores.append(f"{nombre}: {str(e)[:50]}")

    tree_out = ET.ElementTree(root_kml)
    ET.indent(tree_out, space="  ")
    with open("mapa_total.kml", "wb") as f:
        tree_out.write(f, xml_declaration=True, encoding="utf-8")

    msg = f"LISTO ✔ {exitosos}/{len(mapas)} mapas unidos en mapa_total.kml"
    if errores:
        msg += f"\n\nErrores ({len(errores)}):\n" + "\n".join(errores[:10])
        if len(errores) > 10:
            msg += f"\n... y {len(errores)-10} más"
    messagebox.showinfo("Resultado", msg)

frame_btn = tk.Frame(root, padx=10, pady=10)
frame_btn.pack(fill="x")

frame_btn_top = tk.Frame(frame_btn)
frame_btn_top.pack(fill="x", pady=(0, 6))

tk.Button(frame_btn_top, text="📂 Importar Excel", command=importar_excel, bg="#FF9800", fg="white", width=18).pack(side="left", padx=4)
tk.Button(frame_btn_top, text="� Generar por Sesiones", command=importar_por_sesiones, bg="#9C27B0", fg="white", width=22).pack(side="left", padx=4)

frame_btn_bottom = tk.Frame(frame_btn)
frame_btn_bottom.pack(fill="x")

tk.Button(frame_btn_bottom, text="➕ Agregar mapa",  command=agregar,  bg="#4CAF50", fg="white", width=18).pack(side="left", padx=4)
tk.Button(frame_btn_bottom, text="🗑 Eliminar selec.", command=eliminar, bg="#f44336", fg="white", width=18).pack(side="left", padx=4)
tk.Button(frame_btn_bottom, text="⚙ Generar KML",   command=generar,  bg="#2196F3", fg="white", width=18).pack(side="left", padx=4)

root.mainloop()
