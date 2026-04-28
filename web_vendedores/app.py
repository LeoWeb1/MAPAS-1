from flask import Flask, request, jsonify, render_template, send_file
from openpyxl import Workbook, load_workbook
from datetime import datetime
from xml.etree import ElementTree as ET
import requests
import re
import os

app = Flask(__name__)
EXCEL_FILE = "vendedores.xlsx"
KML_FILE = "mapa_unido.kml"

NS = "http://www.opengis.net/kml/2.2"
ET.register_namespace("", NS)

# Colores KML en formato AABBGGRR
COLORES = [
    "ff0000ff",  # Rojo
    "ffff0000",  # Azul
    "ff00ff00",  # Verde
    "ff00ffff",  # Amarillo
    "ff0080ff",  # Naranja
    "ffff00ff",  # Morado
    "ffffff00",  # Celeste
    "ff8080ff",  # Rosa
    "ff00ffaa",  # Verde lima
    "ffff8000",  # Azul cielo
]

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendedores"
        ws.append(["Nombre", "Enlace", "Fecha y Hora"])
        wb.save(EXCEL_FILE)

def guardar_en_excel(nombre, enlace):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([nombre, enlace, datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    wb.save(EXCEL_FILE)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/descargar")
def descargar():
    if not os.path.exists(EXCEL_FILE):
        return "No hay datos aún.", 404
    return send_file(EXCEL_FILE, as_attachment=True)

@app.route("/separar", methods=["POST"])
def separar():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(["--- SEPARADOR ---", "--- NUEVA SESIÓN ---", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    wb.save(EXCEL_FILE)
    return jsonify({"ok": True, "msg": "✂️ Sesión separada. Los siguientes mapas irán en otro grupo."})

@app.route("/limpiar", methods=["POST"])
def limpiar():
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendedores"
    ws.append(["Nombre", "Enlace", "Fecha y Hora"])
    wb.save(EXCEL_FILE)
    return jsonify({"ok": True, "msg": "🗑️ Excel limpiado. Puedes empezar de nuevo."})

@app.route("/guardar", methods=["POST"])
def guardar():
    data = request.get_json()
    nombre = data.get("nombre", "").strip()
    enlace = data.get("enlace", "").strip()

    if not nombre or not enlace:
        return jsonify({"ok": False, "msg": "Nombre y enlace son requeridos."}), 400

    guardar_en_excel(nombre, enlace)
    return jsonify({"ok": True, "msg": f"¡Listo! Mapa de {nombre} guardado."})

if __name__ == "__main__":
    init_excel()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
